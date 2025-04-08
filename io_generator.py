#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
IO点表生成器模块

此模块负责IO点表生成和数据处理，包含以下主要组件：
1. FormFields类 - 定义了表单字段结构
2. IOChannelModels类 - IO通道模型配置类
3. IOChannelCalculator类 - 用于IO通道计算和Excel导出

该模块主要负责IO点表的生成和导出功能，是整个应用的数据处理层
JianDaoYunAPI类已移至api.py模块，实现了关注点分离
"""
import pandas as pd
import os
from typing import List, Dict, Any
from PySide6.QtWidgets import QMessageBox

class FormFields:
    """
    表单字段数据结构，便于后期增删改
    
    该类定义了所有表单字段的结构，包括主表单和子表单字段
    通过嵌套类组织字段，提高代码的可读性和可维护性
    """
    
    # 主表单字段
    class MainForm:
        """
        主表单字段定义
        包含项目基本信息字段
        """
        PROJECT_NAME = "项目名称"
        PROJECT_NUMBER = "项目编号"
        SHENHUA_NUMBER = "深化设计编号"
        CLIENT_NAME = "客户名称"
        STATION = "场站"
        
        @classmethod
        def get_all_fields(cls) -> List[str]:
            """
            获取所有主表单字段
            返回一个包含所有字段名称的列表
            """
            return [
                cls.PROJECT_NAME,
                cls.STATION,
                cls.PROJECT_NUMBER,
                cls.SHENHUA_NUMBER,
                cls.CLIENT_NAME
            ]
    
    # 深化清单子表单字段
    class ShenhuaForm:
        """
        深化清单子表单字段定义
        包含设备详细信息字段
        """
        EQUIPMENT_NAME = "设备名称"
        BRAND = "品牌"
        SPEC_MODEL = "规格型号"
        TECH_PARAMS = "技术参数"
        QUANTITY = "数量"
        UNIT = "单位"
        SUBSYSTEM = "子系统"
        REMARK = "备注"
        TECH_REMARK = "技术备注"
        CONTRACT_STATUS = "合同内外"
        
        @classmethod
        def get_all_fields(cls) -> List[str]:
            """
            获取所有深化清单字段
            返回一个包含所有字段名称的列表
            """
            return [
                cls.EQUIPMENT_NAME,
                cls.BRAND,
                cls.SPEC_MODEL,
                cls.TECH_PARAMS,
                cls.QUANTITY,
                cls.UNIT,
                cls.SUBSYSTEM,
                cls.REMARK,
                cls.TECH_REMARK,
                cls.CONTRACT_STATUS
            ]

# JianDaoYunAPI类已移至api.py模块

class IOChannelModels:
    """
    IO通道模型配置类
    
    封装所有与IO通道和点表相关的配置数据
    提供统一的接口来访问和管理这些数据
    支持动态扩展和修改
    """
    
    # 需要使用BOOL类型地址的字段前缀
    BOOL_TYPE_ADDRESS_FIELDS = [
        "LL报警", "L报警", "H报警", "HH报警", "维护使能开关点位"
    ]
    
    # 需要用户填写的字段（将在导出时高亮显示）
    HIGHLIGHT_FIELDS = [
        "供电类型（有源/无源）", "线制", "位号", "变量名称（HMI）", "变量描述",
        "量程低限", "量程高限", "SLL设定值", "SL设定值", "SH设定值", "SHH设定值"
    ]
    
    @classmethod
    def is_bool_address_field(cls, field_name):
        """
        判断字段是否应该使用BOOL类型地址
        
        Args:
            field_name: 字段名称
            
        Returns:
            bool: 是否应使用BOOL类型地址
        """
        for prefix in cls.BOOL_TYPE_ADDRESS_FIELDS:
            if field_name.startswith(prefix):
                return True
        return False
    
    @classmethod
    def set_highlight_fields(cls, fields: List[str]) -> bool:
        """
        设置需要高亮显示的字段列表
        
        Args:
            fields: 字段名称列表
            
        Returns:
            bool: 操作是否成功
        """
        # 验证所有字段是否存在于IO点表字段中
        io_fields = set(cls.get_io_point_fields())
        for field in fields:
            if field not in io_fields:
                return False
        
        # 更新高亮字段列表
        cls.HIGHLIGHT_FIELDS = fields
        return True
    
    @classmethod
    def add_highlight_field(cls, field: str) -> bool:
        """
        添加一个需要高亮显示的字段
        
        Args:
            field: 字段名称
            
        Returns:
            bool: 操作是否成功
        """
        # 检查字段是否存在
        if field not in cls.get_io_point_fields():
            return False
            
        # 检查字段是否已经在高亮列表中
        if field in cls.HIGHLIGHT_FIELDS:
            return True
            
        # 添加字段到高亮列表
        cls.HIGHLIGHT_FIELDS.append(field)
        return True
    
    @classmethod
    def remove_highlight_field(cls, field: str) -> bool:
        """
        从高亮列表中移除一个字段
        
        Args:
            field: 字段名称
            
        Returns:
            bool: 操作是否成功
        """
        # 检查字段是否在高亮列表中
        if field not in cls.HIGHLIGHT_FIELDS:
            return False
            
        # 从高亮列表中移除字段
        cls.HIGHLIGHT_FIELDS.remove(field)
        return True
    
    @classmethod
    def get_highlight_fields(cls) -> List[str]:
        """
        获取当前需要高亮的字段列表
        
        Returns:
            List[str]: 高亮字段列表
        """
        return cls.HIGHLIGHT_FIELDS.copy()
    
    @classmethod
    def get_model_channel_mapping(cls):
        """
        获取设备型号对应的通道类型与数量
        
        Returns:
            dict: 设备型号映射数据
        """
        return {
            "LK610": {"type": "DI", "channels": 16, "data_type": "BOOL"},  # 16通道DI，布尔类型
            "LK710": {"type": "DO", "channels": 16, "data_type": "BOOL"},  # 16通道DO，布尔类型
            "LK411": {"type": "AI", "channels": 8, "data_type": "REAL"},   # 8通道AI，实数类型
            "LK512": {"type": "AO", "channels": 8, "data_type": "REAL"},   # 8通道AO，实数类型
        }
    
    @classmethod
    def get_io_point_fields(cls):
        """
        获取IO点表的表头字段
        
        Returns:
            list: 表头字段列表
        """
        return [
            "序号", "模块名称", "模块类型", "供电类型（有源/无源）", "线制", "通道位号", "位号", "场站名", 
            "变量名称（HMI）", "变量描述", "数据类型", "读写属性", "保存历史", "掉电保护", 
            "量程低限", "量程高限", "SLL设定值", "SLL设定点位", "SLL设定点位_PLC地址", "SLL设定点位_通讯地址",
            "SL设定值", "SL设定点位", "SL设定点位_PLC地址", "SL设定点位_通讯地址",
            "SH设定值", "SH设定点位", "SH设定点位_PLC地址", "SH设定点位_通讯地址",
            "SHH设定值", "SHH设定点位", "SHH设定点位_PLC地址", "SHH设定点位_通讯地址",
            "LL报警", "LL报警_PLC地址", "LL报警_通讯地址",
            "L报警", "L报警_PLC地址", "L报警_通讯地址",
            "H报警", "H报警_PLC地址", "H报警_通讯地址",
            "HH报警", "HH报警_PLC地址", "HH报警_通讯地址",
            "维护值设定", "维护值设定点位", "维护值设定点位_PLC地址", "维护值设定点位_通讯地址", 
            "维护使能开关点位", "维护使能开关点位_PLC地址", "维护使能开关点位_通讯地址",
            "PLC绝对地址", "上位机通讯地址"
        ]
    
    @classmethod
    def get_field_index(cls, field_name):
        """
        获取字段在列表中的索引
        
        Args:
            field_name: 字段名称
            
        Returns:
            int: 字段索引，不存在则返回-1
        """
        fields = cls.get_io_point_fields()
        try:
            return fields.index(field_name)
        except ValueError:
            return -1
    
    @classmethod
    def add_model(cls, model_key, channel_type, channel_count, data_type):
        """
        添加新的设备型号配置
        注意: 此方法只在当前运行期间有效，不会持久化保存
        
        Args:
            model_key: 设备型号标识
            channel_type: 通道类型
            channel_count: 通道数量
            data_type: 数据类型
            
        Returns:
            bool: 添加是否成功
        """
        # 动态修改类属性
        models = cls.get_model_channel_mapping()
        models[model_key] = {
            "type": channel_type, 
            "channels": channel_count, 
            "data_type": data_type
        }
        return True
    
    @classmethod
    def add_field(cls, field_name, position=None):
        """
        添加新的表头字段
        注意: 此方法只在当前运行期间有效，不会持久化保存
        
        Args:
            field_name: 字段名称
            position: 插入位置，默认添加到末尾
            
        Returns:
            bool: 添加是否成功
        """
        fields = cls.get_io_point_fields()
        if field_name in fields:
            return False  # 字段已存在
            
        if position is None or position >= len(fields):
            fields.append(field_name)
        else:
            fields.insert(position, field_name)
            
        return True

class IOChannelCalculator:
    """
    IO通道计算工具类
    
    负责计算IO通道数量、类型，以及生成Excel报表
    提供基础的通道计算和数据处理功能
    """
    
    # 使用IOChannelModels类获取配置数据
    @classmethod
    def calculate_channels(cls, equipment_list: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """
        根据设备清单计算各类型通道总数及数据类型
        
        Args:
            equipment_list: 设备清单数据列表
        
        Returns:
            dict: 包含各类型通道总数和数据类型的字典 {'AI': {'count': x, 'data_type': 'REAL'}, ...}
        """
        # 初始化结果
        channel_totals = {
            "AI": {"count": 0, "data_type": "REAL"},
            "AO": {"count": 0, "data_type": "REAL"},
            "DI": {"count": 0, "data_type": "BOOL"},
            "DO": {"count": 0, "data_type": "BOOL"}
        }
        
        # 获取设备型号映射
        model_channel_mapping = IOChannelModels.get_model_channel_mapping()
        
        # 处理每个设备
        for equipment in equipment_list:
            # 获取设备规格型号
            spec_model = equipment.get("规格型号", "")
            if not spec_model:
                continue
                
            # 获取数量
            quantity = int(equipment.get("数量", 0))
            if quantity <= 0:
                continue
            
            # 查找匹配的设备型号
            for model_key, channel_info in model_channel_mapping.items():
                if model_key in spec_model:
                    channel_type = channel_info["type"]
                    channel_count = channel_info["channels"]
                    
                    # 计算总通道数并添加到对应类型
                    total_channels = quantity * channel_count
                    channel_totals[channel_type]["count"] += total_channels
                    break
        
        return channel_totals
    
    @classmethod
    def calculate_modbus_address(cls, plc_address: str, data_type: str) -> int:
        """
        根据PLC绝对地址计算上位机通讯地址
        
        Args:
            plc_address: PLC绝对地址，如"%MD100"或"%MX20.0"
            data_type: 数据类型，"REAL"或"BOOL"
            
        Returns:
            int: 计算出的通讯地址
        """
        if data_type == "REAL":
            # 对于REAL类型：=(MID(AE2,4,4)/2)+43001
            # 从%MD100中提取100，然后计算
            md_num = int(plc_address[3:])  # 提取%MD后面的数字
            return (md_num // 2) + 43001
        else:  # BOOL类型
            # 对于BOOL类型：=(MID(AE3,4,2)*8)+RIGHT(AE3,1)+3001
            # 从%MX20.0中提取20和0，然后计算
            parts = plc_address[3:].split('.')  # 分割%MX后面的部分，如"20.0"
            mx_num = int(parts[0])  # 提取前面的数字，如"20"
            bit_num = int(parts[1])  # 提取后面的数字，如"0"
            return (mx_num * 8) + bit_num + 3001
    
    @classmethod
    def export_to_excel(cls, channel_data: Dict[str, Dict[str, Any]], output_path: str = "IO点表.xlsx", equipment_list: List[Dict[str, Any]] = None) -> bool:
        """
        将IO通道数据导出到Excel
        
        Args:
            channel_data: 通道数据字典，包含通道数量和数据类型
            output_path: 输出Excel文件路径
            equipment_list: 设备清单数据列表，用于生成详细设备表
            
        Returns:
            bool: 操作是否成功
        """
        try:
            # 解决文件权限问题 - 使用临时文件
            output_dir = os.path.dirname(output_path)
            temp_file = os.path.join(output_dir, f"temp_{os.path.basename(output_path)}")
            
            # 准备IO点表数据
            io_points = []
            
            # 通道计数器（用于生成通道位号）
            module_counters = {
                "AI": 1,  # 从模块1开始
                "AO": 1,
                "DI": 1,
                "DO": 1
            }
            
            # 序号计数器
            index_counter = 1
            
            # PLC地址计数器 - 主地址
            real_address_counter = 100  # %MD100开始
            bool_address_counter = [20, 0]  # %MX20.0开始
            
            # 附加地址计数器 - 用于额外的字段地址
            extra_real_address_counter = 500  # 从%MD500开始给额外的REAL地址
            extra_bool_address_counter = [100, 0]  # 从%MX100.0开始给额外的BOOL地址
            
            # 机架信息
            rack_count = 1  # 默认为1个机架
            
            # 查找LK117机架信息
            if equipment_list:
                for equipment in equipment_list:
                    spec_model = equipment.get("规格型号", "")
                    if "LK117" in spec_model:
                        rack_count = int(equipment.get("数量", 1))
                        break
            
            # 当前槽位跟踪
            current_rack = 1
            current_slot = 2  # 从2开始，因为1号槽位用于LK232通信模块
            
            # 可用槽位数量（每个机架10个可用槽，第一个槽位用于通信模块）
            available_slots_per_rack = 10
            
            # 获取设备型号映射
            model_channel_mapping = IOChannelModels.get_model_channel_mapping()
            
            if equipment_list:
                # 按照IO类型对设备进行分类
                io_equipment_groups = {
                    "AI": [],
                    "AO": [],
                    "DI": [],
                    "DO": []
                }
                
                # 遍历设备列表进行分类
                for equipment in equipment_list:
                    spec_model = equipment.get("规格型号", "")
                    # 检查是否是IO模块
                    for model_key, channel_info in model_channel_mapping.items():
                        if model_key in spec_model:
                            io_type = channel_info["type"]
                            io_equipment_groups[io_type].append(equipment)
                            break
                
                # 按照AI/AO/DI/DO的顺序遍历处理设备
                for io_type in ["AI", "AO", "DI", "DO"]:
                    for equipment in io_equipment_groups[io_type]:
                        spec_model = equipment.get("规格型号", "")
                        # 获取该设备的通道信息
                        is_io_module = False
                        channels = 0
                        data_type = ""
                        
                        for model_key, channel_info in model_channel_mapping.items():
                            if model_key in spec_model:
                                is_io_module = True
                                io_type = channel_info["type"]  # 这里重新获取io_type是为了保险
                                channels = channel_info["channels"]
                                data_type = channel_info["data_type"]
                                break
                        
                        if is_io_module:
                            quantity = int(equipment.get("数量", 0))
                            equipment_name = equipment.get("设备名称", "未命名设备")
                            station_name = equipment.get("_station", "")
                            
                            # 为每个设备的每个通道创建单独的点表条目
                            for q in range(quantity):
                                # 获取当前模块号
                                module_num = module_counters[io_type]
                                
                                # 计算机架号和槽号
                                # 当前槽位已达到最大值时，移动到下一个机架
                                if current_slot > 11 or current_slot > available_slots_per_rack + 1:
                                    current_rack += 1
                                    current_slot = 2  # 重置为2，跳过第一个槽位
                                    
                                    # 检查是否超出机架数量
                                    if current_rack > rack_count:
                                        error_msg = f"警告：IO模块数量超出了可用机架数量 ({rack_count})，请增加机架数量。"
                                        QMessageBox.warning(None, "机架不足", error_msg)
                                        # 继续使用最后一个机架
                                
                                # 为该模块的每个通道创建条目
                                for ch in range(channels):
                                    # 生成新的通道位号格式（例如：1_1_AO_0）
                                    channel_code = f"{current_rack}_{current_slot}_{io_type}_{ch}"
                                    
                                    # 生成PLC绝对地址
                                    plc_address = ""
                                    if data_type == "REAL":
                                        plc_address = f"%MD{real_address_counter}"
                                        real_address_counter += 4  # REAL类型每个点位加4
                                    else:  # BOOL类型
                                        plc_address = f"%MX{bool_address_counter[0]}.{bool_address_counter[1]}"
                                        # 更新BOOL地址计数器
                                        bool_address_counter[1] += 1
                                        if bool_address_counter[1] > 7:
                                            bool_address_counter[0] += 1
                                            bool_address_counter[1] = 0
                                    
                                    # 计算上位机通讯地址
                                    modbus_address = cls.calculate_modbus_address(plc_address, data_type)
                                    
                                    # 准备该通道的点表数据
                                    point_data = {
                                        "序号": index_counter,
                                        "模块名称": equipment_name,
                                        "模块类型": io_type,
                                        "供电类型（有源/无源）": "/" if io_type == "AO" else "",
                                        "线制": "/" if io_type == "AO" else "",
                                        "通道位号": channel_code,
                                        "位号": "",
                                        "场站名": station_name,
                                        "变量名称（HMI）": "",
                                        "变量描述": "",
                                        "数据类型": data_type,
                                        "读写属性": "R/W",  # 所有点位统一设置为R/W
                                        "保存历史": "是",   # 默认"是"
                                        "掉电保护": "是",   # 默认"是"
                                        "量程低限": "" if data_type == "REAL" else "/",
                                        "量程高限": "" if data_type == "REAL" else "/",
                                        "SLL设定值": "" if data_type == "REAL" else "/",
                                        "SLL设定点位": "",
                                        "SLL设定点位_PLC地址": "",
                                        "SLL设定点位_通讯地址": "",
                                        "SL设定值": "" if data_type == "REAL" else "/",
                                        "SL设定点位": "",
                                        "SL设定点位_PLC地址": "",
                                        "SL设定点位_通讯地址": "",
                                        "SH设定值": "" if data_type == "REAL" else "/",
                                        "SH设定点位": "",
                                        "SH设定点位_PLC地址": "",
                                        "SH设定点位_通讯地址": "",
                                        "SHH设定值": "" if data_type == "REAL" else "/",
                                        "SHH设定点位": "",
                                        "SHH设定点位_PLC地址": "",
                                        "SHH设定点位_通讯地址": "",
                                        "LL报警": "",
                                        "LL报警_PLC地址": "",
                                        "LL报警_通讯地址": "",
                                        "L报警": "",
                                        "L报警_PLC地址": "",
                                        "L报警_通讯地址": "",
                                        "H报警": "",
                                        "H报警_PLC地址": "",
                                        "H报警_通讯地址": "",
                                        "HH报警": "",
                                        "HH报警_PLC地址": "",
                                        "HH报警_通讯地址": "",
                                        "维护值设定": "/",
                                        "维护值设定点位": "",
                                        "维护值设定点位_PLC地址": "",
                                        "维护值设定点位_通讯地址": "",
                                        "维护使能开关点位": "",
                                        "维护使能开关点位_PLC地址": "",
                                        "维护使能开关点位_通讯地址": "",
                                        "PLC绝对地址": plc_address,
                                        "上位机通讯地址": str(modbus_address)
                                    }
                                    
                                    # 添加到点表列表
                                    io_points.append(point_data)
                                    index_counter += 1
                                
                                # 每个设备模块增加模块计数器和槽位计数器
                                module_counters[io_type] += 1
                                current_slot += 1
            
            # 创建Excel写入器
            with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                # 只创建IO点表
                if io_points:
                    # 创建DataFrame并确保列顺序与IO_POINT_FIELDS一致
                    points_df = pd.DataFrame(io_points)
                    
                    # 获取IO点表字段
                    io_point_fields = IOChannelModels.get_io_point_fields()
                    
                    # 确保所有字段都在DataFrame中
                    for field in io_point_fields:
                        if field not in points_df.columns:
                            points_df[field] = ""
                    
                    # 按照IO_POINT_FIELDS中定义的顺序重排列
                    points_df = points_df[io_point_fields]
                    
                    # 导出到Excel
                    points_df.to_excel(writer, sheet_name="IO点表", index=False)
                else:
                    # 没有点位数据时，创建一个空表
                    empty_df = pd.DataFrame(columns=IOChannelModels.get_io_point_fields())
                    empty_df.to_excel(writer, sheet_name="IO点表", index=False)
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets["IO点表"]
                
                # 导入openpyxl的边框样式和填充样式
                from openpyxl.styles import Border, Side, PatternFill, Font
                
                # 定义边框样式（细线边框）
                thin_border = Border(left=Side(style='thin'), 
                                  right=Side(style='thin'), 
                                  top=Side(style='thin'), 
                                  bottom=Side(style='thin'))
                
                # 定义黄色填充样式（用于高亮需要用户填写的字段）
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                # 定义深黄色填充样式（用于高亮表头）
                dark_yellow_fill = PatternFill(start_color="FFDD00", end_color="FFDD00", fill_type="solid")
                # 定义表头字体样式
                header_font = Font(bold=True)
                
                # 获取表头行的单元格范围
                header_row = 1  # Excel中的第一行
                max_col = len(io_point_fields)
                
                # 应用表头字体样式和边框
                for col in range(1, max_col + 1):  # Excel列从1开始
                    cell = worksheet.cell(row=header_row, column=col)
                    cell.font = header_font
                    cell.border = thin_border
                
                # 查找需要高亮的列索引
                highlight_cols = []
                for field in IOChannelModels.get_highlight_fields():
                    col_idx = IOChannelModels.get_field_index(field)
                    if col_idx != -1:  # 确保字段存在
                        highlight_cols.append(col_idx + 1)  # Excel列从1开始
                        # 同时高亮表头（使用深黄色）
                        header_cell = worksheet.cell(row=header_row, column=col_idx + 1)
                        header_cell.fill = dark_yellow_fill
                
                # 查找各关键列的索引
                hmi_col_idx = IOChannelModels.get_field_index("变量名称（HMI）") + 1  # Excel列从1开始
                
                # 各设定点位列的索引
                sll_pos_col_idx = IOChannelModels.get_field_index("SLL设定点位") + 1
                sl_pos_col_idx = IOChannelModels.get_field_index("SL设定点位") + 1
                sh_pos_col_idx = IOChannelModels.get_field_index("SH设定点位") + 1
                shh_pos_col_idx = IOChannelModels.get_field_index("SHH设定点位") + 1
                
                # 各报警点位列的索引
                ll_alarm_col_idx = IOChannelModels.get_field_index("LL报警") + 1
                l_alarm_col_idx = IOChannelModels.get_field_index("L报警") + 1
                h_alarm_col_idx = IOChannelModels.get_field_index("H报警") + 1
                hh_alarm_col_idx = IOChannelModels.get_field_index("HH报警") + 1
                
                # 维护相关列的索引
                maint_val_pos_col_idx = IOChannelModels.get_field_index("维护值设定点位") + 1
                maint_en_pos_col_idx = IOChannelModels.get_field_index("维护使能开关点位") + 1
                
                # 量程上下限的列索引
                low_range_col_idx = IOChannelModels.get_field_index("量程低限") + 1
                high_range_col_idx = IOChannelModels.get_field_index("量程高限") + 1
                
                # 获取数据行数
                data_rows = len(io_points) if io_points else 0
                
                # 为每行添加公式和额外地址，并高亮需要填写的单元格
                for row in range(2, data_rows + 2):  # 从第2行开始（跳过表头）
                    # 获取该行的数据类型
                    data_type_cell = worksheet.cell(row=row, column=IOChannelModels.get_field_index("数据类型") + 1).value
                    
                    # 为所有单元格添加边框
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                    
                    # 针对不同类型字段添加额外地址列
                    if data_type_cell == "REAL":
                        # 为需要添加额外地址的REAL类型字段生成额外地址
                        # SLL设定点位地址
                        sll_addr_col = IOChannelModels.get_field_index("SLL设定点位_PLC地址") + 1
                        sll_comm_col = IOChannelModels.get_field_index("SLL设定点位_通讯地址") + 1
                        extra_plc_addr = f"%MD{extra_real_address_counter}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "REAL")
                        worksheet.cell(row=row, column=sll_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=sll_comm_col).value = extra_comm_addr
                        extra_real_address_counter += 4
                        
                        # SL设定点位地址
                        sl_addr_col = IOChannelModels.get_field_index("SL设定点位_PLC地址") + 1
                        sl_comm_col = IOChannelModels.get_field_index("SL设定点位_通讯地址") + 1
                        extra_plc_addr = f"%MD{extra_real_address_counter}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "REAL")
                        worksheet.cell(row=row, column=sl_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=sl_comm_col).value = extra_comm_addr
                        extra_real_address_counter += 4
                        
                        # SH设定点位地址
                        sh_addr_col = IOChannelModels.get_field_index("SH设定点位_PLC地址") + 1
                        sh_comm_col = IOChannelModels.get_field_index("SH设定点位_通讯地址") + 1
                        extra_plc_addr = f"%MD{extra_real_address_counter}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "REAL")
                        worksheet.cell(row=row, column=sh_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=sh_comm_col).value = extra_comm_addr
                        extra_real_address_counter += 4
                        
                        # SHH设定点位地址
                        shh_addr_col = IOChannelModels.get_field_index("SHH设定点位_PLC地址") + 1
                        shh_comm_col = IOChannelModels.get_field_index("SHH设定点位_通讯地址") + 1
                        extra_plc_addr = f"%MD{extra_real_address_counter}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "REAL")
                        worksheet.cell(row=row, column=shh_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=shh_comm_col).value = extra_comm_addr
                        extra_real_address_counter += 4
                        
                        # 维护值设定点位地址
                        maint_val_addr_col = IOChannelModels.get_field_index("维护值设定点位_PLC地址") + 1
                        maint_val_comm_col = IOChannelModels.get_field_index("维护值设定点位_通讯地址") + 1
                        extra_plc_addr = f"%MD{extra_real_address_counter}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "REAL")
                        worksheet.cell(row=row, column=maint_val_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=maint_val_comm_col).value = extra_comm_addr
                        extra_real_address_counter += 4
                        
                        # 为BOOL类型地址字段添加BOOL类型地址
                        # LL报警地址(BOOL)
                        ll_alarm_addr_col = IOChannelModels.get_field_index("LL报警_PLC地址") + 1
                        ll_alarm_comm_col = IOChannelModels.get_field_index("LL报警_通讯地址") + 1
                        extra_plc_addr = f"%MX{extra_bool_address_counter[0]}.{extra_bool_address_counter[1]}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "BOOL")
                        worksheet.cell(row=row, column=ll_alarm_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=ll_alarm_comm_col).value = extra_comm_addr
                        # 更新BOOL地址计数器
                        extra_bool_address_counter[1] += 1
                        if extra_bool_address_counter[1] > 7:
                            extra_bool_address_counter[0] += 1
                            extra_bool_address_counter[1] = 0
                        
                        # L报警地址(BOOL)
                        l_alarm_addr_col = IOChannelModels.get_field_index("L报警_PLC地址") + 1
                        l_alarm_comm_col = IOChannelModels.get_field_index("L报警_通讯地址") + 1
                        extra_plc_addr = f"%MX{extra_bool_address_counter[0]}.{extra_bool_address_counter[1]}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "BOOL")
                        worksheet.cell(row=row, column=l_alarm_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=l_alarm_comm_col).value = extra_comm_addr
                        # 更新BOOL地址计数器
                        extra_bool_address_counter[1] += 1
                        if extra_bool_address_counter[1] > 7:
                            extra_bool_address_counter[0] += 1
                            extra_bool_address_counter[1] = 0
                        
                        # H报警地址(BOOL)
                        h_alarm_addr_col = IOChannelModels.get_field_index("H报警_PLC地址") + 1
                        h_alarm_comm_col = IOChannelModels.get_field_index("H报警_通讯地址") + 1
                        extra_plc_addr = f"%MX{extra_bool_address_counter[0]}.{extra_bool_address_counter[1]}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "BOOL")
                        worksheet.cell(row=row, column=h_alarm_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=h_alarm_comm_col).value = extra_comm_addr
                        # 更新BOOL地址计数器
                        extra_bool_address_counter[1] += 1
                        if extra_bool_address_counter[1] > 7:
                            extra_bool_address_counter[0] += 1
                            extra_bool_address_counter[1] = 0
                        
                        # HH报警地址(BOOL)
                        hh_alarm_addr_col = IOChannelModels.get_field_index("HH报警_PLC地址") + 1
                        hh_alarm_comm_col = IOChannelModels.get_field_index("HH报警_通讯地址") + 1
                        extra_plc_addr = f"%MX{extra_bool_address_counter[0]}.{extra_bool_address_counter[1]}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "BOOL")
                        worksheet.cell(row=row, column=hh_alarm_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=hh_alarm_comm_col).value = extra_comm_addr
                        # 更新BOOL地址计数器
                        extra_bool_address_counter[1] += 1
                        if extra_bool_address_counter[1] > 7:
                            extra_bool_address_counter[0] += 1
                            extra_bool_address_counter[1] = 0
                        
                        # 维护使能开关点位地址(BOOL)
                        maint_en_addr_col = IOChannelModels.get_field_index("维护使能开关点位_PLC地址") + 1
                        maint_en_comm_col = IOChannelModels.get_field_index("维护使能开关点位_通讯地址") + 1
                        extra_plc_addr = f"%MX{extra_bool_address_counter[0]}.{extra_bool_address_counter[1]}"
                        extra_comm_addr = cls.calculate_modbus_address(extra_plc_addr, "BOOL")
                        worksheet.cell(row=row, column=maint_en_addr_col).value = extra_plc_addr
                        worksheet.cell(row=row, column=maint_en_comm_col).value = extra_comm_addr
                        # 更新BOOL地址计数器
                        extra_bool_address_counter[1] += 1
                        if extra_bool_address_counter[1] > 7:
                            extra_bool_address_counter[0] += 1
                            extra_bool_address_counter[1] = 0
                        
                        # 为REAL类型设置公式字段
                        # 设置SLL设定点位公式: 变量名称（HMI）+ "_LoLoLimit"
                        sll_pos_col_idx = IOChannelModels.get_field_index("SLL设定点位") + 1
                        worksheet.cell(row=row, column=sll_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_LoLoLimit")'
                        
                        # 设置SL设定点位公式: 变量名称（HMI）+ "_LoLimit"
                        sl_pos_col_idx = IOChannelModels.get_field_index("SL设定点位") + 1
                        worksheet.cell(row=row, column=sl_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_LoLimit")'
                        
                        # 设置SH设定点位公式: 变量名称（HMI）+ "_HiLimit"
                        sh_pos_col_idx = IOChannelModels.get_field_index("SH设定点位") + 1
                        worksheet.cell(row=row, column=sh_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_HiLimit")'
                        
                        # 设置SHH设定点位公式: 变量名称（HMI）+ "_HiHiLimit"
                        shh_pos_col_idx = IOChannelModels.get_field_index("SHH设定点位") + 1
                        worksheet.cell(row=row, column=shh_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_HiHiLimit")'
                        
                        # 设置LL报警公式: 变量名称（HMI）+ "_LL"
                        ll_alarm_col_idx = IOChannelModels.get_field_index("LL报警") + 1
                        worksheet.cell(row=row, column=ll_alarm_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_LL")'
                        
                        # 设置L报警公式: 变量名称（HMI）+ "_L"
                        l_alarm_col_idx = IOChannelModels.get_field_index("L报警") + 1
                        worksheet.cell(row=row, column=l_alarm_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_L")'
                        
                        # 设置H报警公式: 变量名称（HMI）+ "_H"
                        h_alarm_col_idx = IOChannelModels.get_field_index("H报警") + 1
                        worksheet.cell(row=row, column=h_alarm_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_H")'
                        
                        # 设置HH报警公式: 变量名称（HMI）+ "_HH"
                        hh_alarm_col_idx = IOChannelModels.get_field_index("HH报警") + 1
                        worksheet.cell(row=row, column=hh_alarm_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_HH")'
                        
                        # 设置维护值设定点位公式: 变量名称（HMI）+ "_whz"
                        maint_val_pos_col_idx = IOChannelModels.get_field_index("维护值设定点位") + 1
                        worksheet.cell(row=row, column=maint_val_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_whz")'
                        
                        # 设置维护使能开关点位公式: 变量名称（HMI）+ "_MAIN_EN"
                        maint_en_pos_col_idx = IOChannelModels.get_field_index("维护使能开关点位") + 1
                        worksheet.cell(row=row, column=maint_en_pos_col_idx).value = f'=IF(ISBLANK({chr(64+hmi_col_idx)}{row}),"",{chr(64+hmi_col_idx)}{row}&"_MAIN_EN")'
                        
                        # 设置量程低限和量程高限
                        low_range_col_idx = IOChannelModels.get_field_index("量程低限") + 1
                        high_range_col_idx = IOChannelModels.get_field_index("量程高限") + 1
                        # 量程低限和高限可以自行填写，这里不设置默认公式
                        
                        # 设置维护值设定为"/"
                        maint_val_col_idx = IOChannelModels.get_field_index("维护值设定") + 1
                        worksheet.cell(row=row, column=maint_val_col_idx).value = "/"
                    else:
                        # BOOL类型不需要额外地址和公式，所有相关列设置为"/"
                        # 获取所有相关列的索引
                        sll_pos_col_idx = IOChannelModels.get_field_index("SLL设定点位") + 1
                        sll_addr_col = IOChannelModels.get_field_index("SLL设定点位_PLC地址") + 1
                        sll_comm_col = IOChannelModels.get_field_index("SLL设定点位_通讯地址") + 1
                        
                        sl_pos_col_idx = IOChannelModels.get_field_index("SL设定点位") + 1
                        sl_addr_col = IOChannelModels.get_field_index("SL设定点位_PLC地址") + 1
                        sl_comm_col = IOChannelModels.get_field_index("SL设定点位_通讯地址") + 1
                        
                        sh_pos_col_idx = IOChannelModels.get_field_index("SH设定点位") + 1
                        sh_addr_col = IOChannelModels.get_field_index("SH设定点位_PLC地址") + 1
                        sh_comm_col = IOChannelModels.get_field_index("SH设定点位_通讯地址") + 1
                        
                        shh_pos_col_idx = IOChannelModels.get_field_index("SHH设定点位") + 1
                        shh_addr_col = IOChannelModels.get_field_index("SHH设定点位_PLC地址") + 1
                        shh_comm_col = IOChannelModels.get_field_index("SHH设定点位_通讯地址") + 1
                        
                        ll_alarm_col_idx = IOChannelModels.get_field_index("LL报警") + 1
                        ll_alarm_addr_col = IOChannelModels.get_field_index("LL报警_PLC地址") + 1
                        ll_alarm_comm_col = IOChannelModels.get_field_index("LL报警_通讯地址") + 1
                        
                        l_alarm_col_idx = IOChannelModels.get_field_index("L报警") + 1
                        l_alarm_addr_col = IOChannelModels.get_field_index("L报警_PLC地址") + 1
                        l_alarm_comm_col = IOChannelModels.get_field_index("L报警_通讯地址") + 1
                        
                        h_alarm_col_idx = IOChannelModels.get_field_index("H报警") + 1
                        h_alarm_addr_col = IOChannelModels.get_field_index("H报警_PLC地址") + 1
                        h_alarm_comm_col = IOChannelModels.get_field_index("H报警_通讯地址") + 1
                        
                        hh_alarm_col_idx = IOChannelModels.get_field_index("HH报警") + 1
                        hh_alarm_addr_col = IOChannelModels.get_field_index("HH报警_PLC地址") + 1
                        hh_alarm_comm_col = IOChannelModels.get_field_index("HH报警_通讯地址") + 1
                        
                        maint_val_pos_col_idx = IOChannelModels.get_field_index("维护值设定点位") + 1
                        maint_val_addr_col = IOChannelModels.get_field_index("维护值设定点位_PLC地址") + 1
                        maint_val_comm_col = IOChannelModels.get_field_index("维护值设定点位_通讯地址") + 1
                        
                        maint_en_pos_col_idx = IOChannelModels.get_field_index("维护使能开关点位") + 1
                        maint_en_addr_col = IOChannelModels.get_field_index("维护使能开关点位_PLC地址") + 1
                        maint_en_comm_col = IOChannelModels.get_field_index("维护使能开关点位_通讯地址") + 1
                        
                        low_range_col_idx = IOChannelModels.get_field_index("量程低限") + 1
                        high_range_col_idx = IOChannelModels.get_field_index("量程高限") + 1
                        
                        # 设置所有公式字段为"/"
                        # 设定值也需要设置为"/"
                        sll_val_col_idx = IOChannelModels.get_field_index("SLL设定值") + 1
                        sl_val_col_idx = IOChannelModels.get_field_index("SL设定值") + 1
                        sh_val_col_idx = IOChannelModels.get_field_index("SH设定值") + 1
                        shh_val_col_idx = IOChannelModels.get_field_index("SHH设定值") + 1
                        
                        # 设置设定值为"/"
                        worksheet.cell(row=row, column=sll_val_col_idx).value = "/"
                        worksheet.cell(row=row, column=sl_val_col_idx).value = "/"
                        worksheet.cell(row=row, column=sh_val_col_idx).value = "/"
                        worksheet.cell(row=row, column=shh_val_col_idx).value = "/"
                        
                        # 设置设定点位和地址为"/"
                        worksheet.cell(row=row, column=sll_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=sll_addr_col).value = "/"
                        worksheet.cell(row=row, column=sll_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=sl_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=sl_addr_col).value = "/"
                        worksheet.cell(row=row, column=sl_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=sh_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=sh_addr_col).value = "/"
                        worksheet.cell(row=row, column=sh_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=shh_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=shh_addr_col).value = "/"
                        worksheet.cell(row=row, column=shh_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=ll_alarm_col_idx).value = "/"
                        worksheet.cell(row=row, column=ll_alarm_addr_col).value = "/"
                        worksheet.cell(row=row, column=ll_alarm_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=l_alarm_col_idx).value = "/"
                        worksheet.cell(row=row, column=l_alarm_addr_col).value = "/"
                        worksheet.cell(row=row, column=l_alarm_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=h_alarm_col_idx).value = "/"
                        worksheet.cell(row=row, column=h_alarm_addr_col).value = "/"
                        worksheet.cell(row=row, column=h_alarm_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=hh_alarm_col_idx).value = "/"
                        worksheet.cell(row=row, column=hh_alarm_addr_col).value = "/"
                        worksheet.cell(row=row, column=hh_alarm_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=maint_val_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=maint_val_addr_col).value = "/"
                        worksheet.cell(row=row, column=maint_val_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=maint_en_pos_col_idx).value = "/"
                        worksheet.cell(row=row, column=maint_en_addr_col).value = "/"
                        worksheet.cell(row=row, column=maint_en_comm_col).value = "/"
                        
                        worksheet.cell(row=row, column=low_range_col_idx).value = "/"
                        worksheet.cell(row=row, column=high_range_col_idx).value = "/"
                        
                        # 设置维护值设定为"/"
                        maint_val_col_idx = IOChannelModels.get_field_index("维护值设定") + 1
                        worksheet.cell(row=row, column=maint_val_col_idx).value = "/"
                    
                    # 在所有值设置完成后，再进行高亮标记
                    # 为需要高亮的列添加黄色背景
                    for col in highlight_cols:
                        cell = worksheet.cell(row=row, column=col)
                        # 判断单元格是否应该高亮：只要值为"/"就不标黄，无需区分模块类型
                        if cell.value != "/":  # 只有当值不是"/"时才标黄
                            cell.fill = yellow_fill
            
            # 保存成功后，移动临时文件到目标位置
            try:
                # 如果目标文件已存在，尝试删除
                if os.path.exists(output_path):
                    os.remove(output_path)
                
                # 重命名临时文件
                os.rename(temp_file, output_path)
                
                return True
            except Exception as e:
                error_msg = f"移动临时文件失败: {str(e)}\n临时文件位置: {temp_file}"
                QMessageBox.critical(None, "文件操作错误", error_msg)
                # 临时文件创建成功，但无法重命名，返回临时文件路径
                return False
            
        except Exception as e:
            QMessageBox.critical(None, "导出错误", f"导出Excel失败: {str(e)}")
            return False 